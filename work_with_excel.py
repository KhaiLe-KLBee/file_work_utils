from openpyxl import load_workbook

def get_value_excel(file_path, sheetName, cellName):
    wb = load_workbook(filename = file_path)
    sheet = wb[sheetName]
    wb.close();
    return sheet[cellName].value
    
def update_value_excel(file_path, sheetName, cellName, value):
    wb = load_workbook(filename = file_path)
    sheet = wb[sheetName]
    sheet[cellName].value = value
    wb.close();
    wb.save(file_path);

if __name__ == '__main__':
    filename = 'test.xlsx'
    # result = get_value_excel(filename, 'A1')
    # print (result)
    update_value_excel(filename, 'Sheet 1', 'A2', '21')